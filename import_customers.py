import argparse
import re
from datetime import datetime
from pathlib import Path

from app import app
from db_neon import get_neon_db as get_db
from psycopg2.extras import execute_values
from utils.helpers import normalize_number_plate, normalize_phone


DEFAULT_EXCEL_PATH = Path("Customer LIst") / "Customer list all.xlsx"


def _load_rows(path):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise SystemExit("openpyxl is required. Install it with: pip install openpyxl") from error

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value or "").strip().lower() for value in rows[0]]
    return [
        {
            headers[index] if index < len(headers) and headers[index] else f"column_{index + 1}": value
            for index, value in enumerate(row)
        }
        for row in rows[1:]
    ]


def _pick(row, candidates):
    normalized = {
        re.sub(r"[^a-z0-9]", "", key.lower()): value
        for key, value in row.items()
    }
    for candidate in candidates:
        key = re.sub(r"[^a-z0-9]", "", candidate.lower())
        value = normalized.get(key)
        if value is not None and str(value).strip():
            if isinstance(value, float) and value == int(value):
                return str(int(value))
            return str(value).strip()
    return ""


def _normalize_customer_id(value):
    customer_id = (value or "").strip().upper()
    return customer_id if re.fullmatch(r"CUST\d+", customer_id) else ""


def _next_customer_id_start(cursor, imported_ids=None):
    imported_ids = imported_ids or set()
    cursor.execute("SELECT id FROM customers WHERE id LIKE 'CUST%'")
    highest = 1000
    for row in cursor.fetchall():
        match = re.search(r"(\d+)$", row[0] or "")
        if match:
            highest = max(highest, int(match.group(1)))
    for customer_id in imported_ids:
        match = re.search(r"(\d+)$", customer_id or "")
        if match:
            highest = max(highest, int(match.group(1)))
    return highest + 1


def _load_existing_customer_maps(cursor):
    cursor.execute("SELECT id, phone FROM customers")
    rows = cursor.fetchall()
    by_id = {row[0]: row[1] for row in rows}
    by_phone = {row[1]: row[0] for row in rows if row[1]}
    return by_id, by_phone


def _load_existing_vehicle_plates(cursor):
    cursor.execute("SELECT number_plate FROM customer_vehicles")
    return {row[0] for row in cursor.fetchall()}


def _ensure_tables(cursor):
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS customers (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            phone TEXT,
            vehicle TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_customers_phone_unique
        ON customers(phone)
        WHERE phone IS NOT NULL AND phone <> ''
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS customer_vehicles (
            id SERIAL PRIMARY KEY,
            customer_id TEXT NOT NULL,
            number_plate TEXT NOT NULL,
            brand_model TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (customer_id) REFERENCES customers(id) ON DELETE CASCADE,
            UNIQUE (number_plate)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS vehicles (
            plate_number TEXT PRIMARY KEY,
            customer_id TEXT NOT NULL,
            brand TEXT DEFAULT '',
            model TEXT DEFAULT '',
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        )
    """)


def import_customers(path, dry_run=False):
    rows = _load_rows(path)
    customers_seen = set()
    vehicles_seen = set()
    customer_rows = []
    customer_vehicle_rows = []
    legacy_vehicle_rows = []
    skipped = 0
    imported_ids = {
        customer_id
        for row in rows
        if (customer_id := _normalize_customer_id(_pick(row, ("customer id", "customer_id", "cust id", "id"))))
    }

    if dry_run:
        for row in rows:
            phone = normalize_phone(_pick(row, ("c_phone", "cphone", "phone", "mobile", "mobile no", "contact", "contact no", "phone number")))
            plate = normalize_number_plate(_pick(row, ("vehi_num", "vehinum", "number plate", "vehicle number", "vehicle no", "registration number", "reg no", "plate")))
            customer_id = _normalize_customer_id(_pick(row, ("customer id", "customer_id", "cust id", "id")))

            if len(phone) != 10 or not plate:
                skipped += 1
                continue

            customers_seen.add(customer_id or phone)
            vehicles_seen.add(plate)

        print(f"Dry run - Customers parsed: {len(customers_seen)}")
        print(f"Dry run - Vehicles parsed: {len(vehicles_seen)}")
        print(f"Rows skipped: {skipped}")
        return

    db = get_db()
    cursor = db.cursor()

    try:
        _ensure_tables(cursor)
        id_counter = _next_customer_id_start(cursor, imported_ids)
        id_to_phone, phone_to_customer_id = _load_existing_customer_maps(cursor)
        existing_vehicle_plates = _load_existing_vehicle_plates(cursor)

        for row in rows:
            phone = normalize_phone(_pick(row, ("c_phone", "cphone", "phone", "mobile", "mobile no", "contact", "contact no", "phone number")))
            plate = normalize_number_plate(_pick(row, ("vehi_num", "vehinum", "number plate", "vehicle number", "vehicle no", "registration number", "reg no", "plate")))
            name = _pick(row, ("c_name", "cname", "name", "customer name", "owner name")) or "Guest Customer"
            brand_model = _pick(row, ("vehi_brand", "vehibrand", "brand model", "brand/model", "vehi_name", "vehiname", "vehicle model", "model", "bike model", "vehicle"))
            imported_customer_id = _normalize_customer_id(_pick(row, ("customer id", "customer_id", "cust id", "id")))

            if len(phone) != 10 or not plate:
                skipped += 1
                continue

            existing_customer_id = phone_to_customer_id.get(phone)
            if existing_customer_id:
                customer_id = existing_customer_id
            elif imported_customer_id and imported_customer_id in id_to_phone:
                customer_id = imported_customer_id
            else:
                customer_id = imported_customer_id or f"CUST{id_counter}"
                id_counter += 1
                phone_to_customer_id[phone] = customer_id
                id_to_phone[customer_id] = phone
                customer_rows.append((customer_id, name, phone, plate, datetime.now()))

            customers_seen.add(customer_id)

            if plate not in existing_vehicle_plates:
                existing_vehicle_plates.add(plate)
                vehicles_seen.add(plate)
                customer_vehicle_rows.append((customer_id, plate, brand_model, datetime.now()))

            brand, model = (brand_model.split(None, 1) + [""])[:2] if brand_model else ("", "")
            legacy_vehicle_rows.append((plate, customer_id, brand, model))

        if customer_rows:
            execute_values(
                cursor,
                """
                INSERT INTO customers (id, name, phone, vehicle, created_at)
                VALUES %s
                ON CONFLICT (id) DO UPDATE
                SET name = EXCLUDED.name,
                    phone = COALESCE(NULLIF(EXCLUDED.phone, ''), customers.phone),
                    vehicle = COALESCE(NULLIF(EXCLUDED.vehicle, ''), customers.vehicle)
                """,
                customer_rows,
                page_size=1000,
            )

        if customer_vehicle_rows:
            execute_values(
                cursor,
                """
                INSERT INTO customer_vehicles (customer_id, number_plate, brand_model, created_at)
                VALUES %s
                ON CONFLICT DO NOTHING
                """,
                customer_vehicle_rows,
                page_size=1000,
            )

        if legacy_vehicle_rows:
            execute_values(
                cursor,
                """
                INSERT INTO vehicles (plate_number, customer_id, brand, model)
                VALUES %s
                ON CONFLICT DO NOTHING
                """,
                legacy_vehicle_rows,
                page_size=1000,
            )

        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        cursor.close()

    print(f"Customers imported/found: {len(customers_seen)}")
    print(f"Vehicles imported: {len(vehicles_seen)}")
    print(f"Rows skipped: {skipped}")


def main():
    parser = argparse.ArgumentParser(description="Import customer and vehicle rows from Excel.")
    parser.add_argument("path", nargs="?", default=str(DEFAULT_EXCEL_PATH), help="Path to .xlsx file")
    parser.add_argument("--dry-run", action="store_true", help="Parse and report counts without writing rows")
    args = parser.parse_args()
    path = Path(args.path)
    if not path.exists():
        raise SystemExit(f"Excel file not found: {path}")

    with app.app_context():
        import_customers(path, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
